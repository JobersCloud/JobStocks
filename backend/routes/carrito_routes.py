# ============================================================
#      ██╗ ██████╗ ██████╗ ███████╗██████╗ ███████╗
#      ██║██╔═══██╗██╔══██╗██╔════╝██╔══██╗██╔════╝
#      ██║██║   ██║██████╔╝█████╗  ██████╔╝███████╗
# ██   ██║██║   ██║██╔══██╗██╔══╝  ██╔══██╗╚════██║
# ╚█████╔╝╚██████╔╝██████╔╝███████╗██║  ██║███████║
#  ╚════╝  ╚═════╝ ╚═════╝ ╚══════╝╚═╝  ╚═╝╚══════╝
#
#                ──  Jobers - Iaucejo  ──
#
# Autor : iaucejo
# Fecha : 2026-01-08
# ============================================================

# ============================================
# ARCHIVO: routes/carrito_routes.py
# ============================================
from flask import Blueprint, jsonify, request, session
from flask_login import login_required, current_user
from utils.auth import csrf_required
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime
import io
import smtplib
import traceback
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders
from email.utils import formatdate
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models.email_config_model import EmailConfigModel
from models.propuesta_model import PropuestaModel
from models.imagen_model import ImagenModel
from models.cliente_model import ClienteModel
import base64

carrito_bp = Blueprint('carrito', __name__, url_prefix='/api/carrito')

# Ya no necesitamos EMAIL_CONFIG hardcodeado, se obtiene de la BD

@carrito_bp.route('', methods=['GET'])
@login_required
def get_carrito():
    """
    Obtener el carrito del usuario actual
    ---
    tags:
      - Carrito
    security:
      - cookieAuth: []
    responses:
      200:
        description: Lista de items en el carrito
        schema:
          type: array
          items:
            type: object
            properties:
              codigo:
                type: string
              descripcion:
                type: string
              formato:
                type: string
              calidad:
                type: string
              color:
                type: string
              existencias:
                type: number
              cantidad_solicitada:
                type: number
              unidad:
                type: string
      401:
        description: No autenticado
    """
    if not current_user.is_authenticated:
        return jsonify({"error": "No autenticado"}), 401
    
    # Obtener carrito de la sesión
    carrito = session.get('carrito', [])
    return jsonify(carrito), 200

@carrito_bp.route('/add', methods=['POST'])
@login_required
@csrf_required
def add_to_carrito():
    """
    Agregar item al carrito
    ---
    tags:
      - Carrito
    security:
      - cookieAuth: []
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            codigo:
              type: string
            descripcion:
              type: string
            formato:
              type: string
            calidad:
              type: string
            color:
              type: string
            tono:
              type: string
            calibre:
              type: string
            existencias:
              type: number
            unidad:
              type: string
            cantidad_solicitada:
              type: number
    responses:
      200:
        description: Item agregado
      401:
        description: No autenticado
    """
    if not current_user.is_authenticated:
        return jsonify({"error": "No autenticado"}), 401
    
    data = request.json
    
    # Obtener o crear carrito
    if 'carrito' not in session:
        session['carrito'] = []
    
    carrito = session['carrito']
    
    # Verificar si el item ya existe (por codigo, calidad, tono, calibre, pallet, caja)
    item_existente = None
    for item in carrito:
        if (item['codigo'] == data['codigo'] and
            item.get('calidad', '') == data.get('calidad', '') and
            item.get('tono', '') == data.get('tono', '') and
            item.get('calibre', '') == data.get('calibre', '') and
            item.get('pallet', '') == data.get('pallet', '') and
            item.get('caja', '') == data.get('caja', '')):
            item_existente = item
            break

    if item_existente:
        # Actualizar cantidad
        item_existente['cantidad_solicitada'] = data.get('cantidad_solicitada', 0)
    else:
        # Agregar nuevo item
        carrito.append({
            'codigo': data['codigo'],
            'descripcion': data['descripcion'],
            'formato': data.get('formato', ''),
            'calidad': data.get('calidad', ''),
            'color': data.get('color', ''),
            'tono': data.get('tono', ''),
            'calibre': data.get('calibre', ''),
            'pallet': data.get('pallet', ''),
            'caja': data.get('caja', ''),
            'existencias': data.get('existencias', 0),
            'unidad': data.get('unidad', ''),
            'cantidad_solicitada': data.get('cantidad_solicitada', 0)
        })
    
    session['carrito'] = carrito
    session.modified = True
    
    return jsonify({
        "message": "Item agregado al carrito",
        "carrito": carrito
    }), 200

@carrito_bp.route('/remove/<int:index>', methods=['DELETE'])
@login_required
@csrf_required
def remove_from_carrito(index):
    """
    Eliminar item del carrito por índice
    ---
    tags:
      - Carrito
    security:
      - cookieAuth: []
    parameters:
      - name: index
        in: path
        type: integer
        required: true
        description: Índice del item en el carrito
    responses:
      200:
        description: Item eliminado
      400:
        description: Índice inválido
      401:
        description: No autenticado
    """
    if not current_user.is_authenticated:
        return jsonify({"error": "No autenticado"}), 401

    carrito = session.get('carrito', [])

    if index < 0 or index >= len(carrito):
        return jsonify({"error": "Índice inválido"}), 400

    carrito.pop(index)

    session['carrito'] = carrito
    session.modified = True

    return jsonify({
        "message": "Item eliminado del carrito",
        "carrito": carrito
    }), 200

@carrito_bp.route('/clear', methods=['DELETE'])
@login_required
@csrf_required
def clear_carrito():
    """
    Vaciar el carrito
    ---
    tags:
      - Carrito
    security:
      - cookieAuth: []
    responses:
      200:
        description: Carrito vaciado
      401:
        description: No autenticado
    """
    if not current_user.is_authenticated:
        return jsonify({"error": "No autenticado"}), 401
    
    session['carrito'] = []
    session.modified = True
    
    return jsonify({"message": "Carrito vaciado"}), 200

@carrito_bp.route('/enviar', methods=['POST'])
@login_required
@csrf_required
def enviar_carrito():
    """
    Generar PDF y enviarlo por email
    ---
    tags:
      - Carrito
    security:
      - cookieAuth: []
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            comentarios:
              type: string
              description: Comentarios adicionales del usuario
    responses:
      200:
        description: Email enviado correctamente
      401:
        description: No autenticado
      500:
        description: Error al enviar email
    """
    if not current_user.is_authenticated:
        return jsonify({"error": "No autenticado"}), 401
    
    carrito = session.get('carrito', [])
    
    if not carrito:
        return jsonify({"error": "El carrito está vacío"}), 400
    
    data = request.json
    referencia = data.get('referencia', '')
    comentarios = data.get('comentarios', '')
    empresa_id = data.get('empresa_id', '1')  # Multi-empresa support
    enviar_copia = data.get('enviar_copia', False)
    cliente_id = data.get('cliente_id', '')
    firma_base64 = data.get('firma', None)  # Firma digital en base64
    usuario = current_user.full_name or current_user.username
    email_usuario = current_user.email if enviar_copia else None

    # Validar cliente_id obligatorio
    if not cliente_id:
        return jsonify({"error": "Debe seleccionar un cliente"}), 400

    # Obtener datos del cliente
    cliente_info = ClienteModel.get_by_codigo(cliente_id, empresa_id)
    if not cliente_info:
        return jsonify({"error": f"Cliente no encontrado: {cliente_id}"}), 400

    try:
        # Generar PDF (con datos del cliente y firma si existe)
        pdf_buffer = generar_pdf_carrito(carrito, usuario, comentarios, referencia, cliente_info, firma_base64)

        # Generar Excel
        excel_buffer = generar_excel_carrito(carrito, usuario, comentarios, referencia)

        # Enviar email (con CC al usuario si está marcado, datos del cliente y firma)
        enviar_email_con_pdf(pdf_buffer, usuario, carrito, comentarios, empresa_id, email_usuario, referencia, excel_buffer, cliente_info, firma_base64)

        # Guardar propuesta en BD (solo si el email se envio correctamente)
        propuesta_id = PropuestaModel.crear_propuesta(
            user_id=current_user.id,
            carrito=carrito,
            comentarios=comentarios,
            empresa_id=empresa_id,
            referencia=referencia,
            cliente_id=cliente_id
        )
        print(f"8️⃣ Propuesta guardada en BD con ID: {propuesta_id}")

        # Limpiar carrito después de enviar
        session['carrito'] = []
        session.modified = True

        return jsonify({
            "message": "Solicitud enviada correctamente por email",
            "propuesta_id": propuesta_id
        }), 200
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print("=" * 60)
        print("ERROR AL ENVIAR EMAIL:")
        print(error_detail)
        print("=" * 60)
        return jsonify({
            "error": f"Error al enviar la solicitud: {str(e)}",
            "detail": error_detail
        }), 500

# ==================== FUNCIONES AUXILIARES ====================

def generar_pdf_carrito(carrito, usuario, comentarios, referencia="", cliente_info=None, firma_base64=None):
    """Genera un PDF con los items del carrito incluyendo imágenes y firma"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)

    story = []
    styles = getSampleStyleSheet()

    # Estilo personalizado para el título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        alignment=TA_CENTER,
        spaceAfter=30
    )

    # Título
    story.append(Paragraph("Solicitud de Stock", title_style))
    story.append(Spacer(1, 0.3*inch))

    # Información general
    info_style = styles['Normal']
    story.append(Paragraph(f"<b>Usuario:</b> {usuario}", info_style))
    story.append(Paragraph(f"<b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", info_style))
    if cliente_info:
        cliente_texto = f"{cliente_info.get('codigo', '')} - {cliente_info.get('razon', '')}"
        story.append(Paragraph(f"<b>Cliente:</b> {cliente_texto}", info_style))
    if referencia:
        story.append(Paragraph(f"<b>Referencia:</b> {referencia}", info_style))
    story.append(Paragraph(f"<b>Total de items:</b> {len(carrito)}", info_style))
    story.append(Spacer(1, 0.3*inch))

    # Obtener thumbnails de los productos
    codigos = [item['codigo'] for item in carrito]
    thumbnails = ImagenModel.get_thumbnails_batch(codigos)

    # Tabla de productos con imágenes
    data = [['Img', 'Código', 'Descripción', 'Formato', 'Cal.', 'Tono', 'Calib.', 'Stock', 'Solicitado']]

    for item in carrito:
        codigo = item['codigo'].strip() if item['codigo'] else item['codigo']

        # Crear imagen para la celda
        img_cell = '-'
        if codigo in thumbnails:
            try:
                img_data = base64.b64decode(thumbnails[codigo])
                img_buffer = io.BytesIO(img_data)
                img_cell = Image(img_buffer, width=0.4*inch, height=0.4*inch)
            except Exception:
                img_cell = '-'

        data.append([
            img_cell,
            item['codigo'],
            item['descripcion'][:25] + '...' if len(item['descripcion']) > 25 else item['descripcion'],
            item.get('formato', '-'),
            item.get('calidad', '-'),
            item.get('tono', '-'),
            item.get('calibre', '-'),
            f"{item['existencias']:.2f}",
            f"{item['cantidad_solicitada']:.2f}"
        ])

    table = Table(data, colWidths=[0.5*inch, 0.9*inch, 1.6*inch, 0.7*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.8*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')])
    ]))

    story.append(table)
    story.append(Spacer(1, 0.3*inch))

    # Comentarios
    if comentarios:
        story.append(Paragraph("<b>Comentarios:</b>", styles['Heading3']))
        story.append(Paragraph(comentarios, styles['Normal']))

    # Firma digital (si existe)
    if firma_base64:
        try:
            # La firma viene como data:image/png;base64,xxxxx
            if firma_base64.startswith('data:'):
                firma_data = firma_base64.split(',')[1]
            else:
                firma_data = firma_base64

            firma_bytes = base64.b64decode(firma_data)
            firma_buffer = io.BytesIO(firma_bytes)
            story.append(Spacer(1, 0.3*inch))
            story.append(Paragraph("<b>Firma:</b>", styles['Heading3']))
            story.append(Spacer(1, 0.1*inch))
            firma_img = Image(firma_buffer, width=2.5*inch, height=0.9*inch)
            story.append(firma_img)
        except Exception as e:
            print(f"⚠️ Error al procesar firma en PDF: {e}")

    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_excel_carrito(carrito, usuario, comentarios="", referencia=""):
    """Genera un archivo Excel con los items del carrito"""
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Propuesta"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Información de cabecera
    ws['A1'] = "SOLICITUD DE STOCK"
    ws['A1'].font = Font(bold=True, size=14, color="667eea")
    ws.merge_cells('A1:J1')

    ws['A2'] = f"Usuario: {usuario}"
    ws['A3'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if referencia:
        ws['A4'] = f"Referencia: {referencia}"
        start_row = 6
    else:
        start_row = 5

    # Encabezados de la tabla
    headers = ['Código', 'Descripción', 'Formato', 'Calidad', 'Tono', 'Calibre', 'Unidad', 'Pallet', 'Caja', 'Cantidad']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos de los productos
    for row_num, item in enumerate(carrito, start_row + 1):
        ws.cell(row=row_num, column=1, value=item.get('codigo', '')).border = thin_border
        ws.cell(row=row_num, column=2, value=item.get('descripcion', '')).border = thin_border
        ws.cell(row=row_num, column=3, value=item.get('formato', '')).border = thin_border
        ws.cell(row=row_num, column=4, value=item.get('calidad', '')).border = thin_border
        ws.cell(row=row_num, column=5, value=item.get('tono', '')).border = thin_border
        ws.cell(row=row_num, column=6, value=item.get('calibre', '')).border = thin_border
        ws.cell(row=row_num, column=7, value=item.get('unidad', '')).border = thin_border
        ws.cell(row=row_num, column=8, value=item.get('pallet', '')).border = thin_border
        ws.cell(row=row_num, column=9, value=item.get('caja', '')).border = thin_border

        cantidad_cell = ws.cell(row=row_num, column=10, value=item.get('cantidad_solicitada', 0))
        cantidad_cell.border = thin_border
        cantidad_cell.alignment = number_alignment

    # Comentarios al final
    if comentarios:
        last_row = start_row + len(carrito) + 2
        ws.cell(row=last_row, column=1, value="Comentarios:").font = Font(bold=True)
        ws.cell(row=last_row + 1, column=1, value=comentarios)
        ws.merge_cells(f'A{last_row + 1}:J{last_row + 1}')

    # Ajustar anchos de columna
    column_widths = [15, 40, 12, 10, 8, 10, 8, 10, 10, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def enviar_email_con_pdf(pdf_buffer, usuario, carrito, comentarios="", empresa_id="1", email_copia=None, referencia="", excel_buffer=None, cliente_info=None, firma_base64=None):
    """Envía el PDF y Excel por email usando configuración de la BD con imágenes CID y firma"""

    print("\n" + "=" * 60)
    print(f"📧 INICIANDO ENVÍO DE EMAIL (Empresa: {empresa_id})")
    if email_copia:
        print(f"📋 CC habilitado para: {email_copia}")
    print("=" * 60)

    # Obtener configuración desde la base de datos
    print("1️⃣ Obteniendo configuración de email desde BD...")
    email_config = EmailConfigModel.get_active_config(empresa_id)

    if not email_config:
        error_msg = "No hay configuración de email activa en la base de datos"
        print(f"❌ ERROR: {error_msg}")
        print("   Solución: Ejecuta este SQL:")
        print("   UPDATE email_config SET activo = 1 WHERE id = 1;")
        print("=" * 60 + "\n")
        raise Exception(error_msg)

    print(f"✅ Configuración obtenida: {email_config['nombre_configuracion']}")
    print(f"   - Servidor SMTP: {email_config['smtp_server']}:{email_config['smtp_port']}")
    print(f"   - De: {email_config['email_from']}")
    print(f"   - Para: {email_config['email_to']}")

    # Crear mensaje con soporte para imágenes embebidas (mixed para adjuntos)
    print("\n2️⃣ Creando mensaje de email...")
    msg = MIMEMultipart('mixed')
    msg['From'] = email_config['email_from']
    msg['To'] = email_config['email_to']
    if email_copia:
        msg['Cc'] = email_copia
    msg['Subject'] = f"Solicitud de Stock - {usuario} - {datetime.now().strftime('%d/%m/%Y')}"
    msg['Date'] = formatdate(localtime=True)
    print(f"✅ Asunto: {msg['Subject']}")
    print(f"✅ Fecha: {msg['Date']}")
    if email_copia:
        print(f"✅ CC: {email_copia}")

    # Cuerpo del email con tabla HTML de productos e imágenes embebidas (CID)
    print("\n3️⃣ Construyendo cuerpo del email con tabla de productos...")

    # Obtener thumbnails de todos los productos del carrito
    codigos = [item['codigo'] for item in carrito]
    print(f"   Obteniendo thumbnails para {len(codigos)} productos...")
    thumbnails = ImagenModel.get_thumbnails_batch(codigos)
    print(f"   ✅ Thumbnails obtenidos: {len(thumbnails)} de {len(codigos)}")

    # Construir filas de la tabla HTML con referencias CID para imágenes
    filas_productos = ""
    imagenes_adjuntar = {}  # {cid: imagen_base64}
    for item in carrito:
        codigo = item['codigo'].strip() if item['codigo'] else item['codigo']
        # Crear celda de imagen
        if codigo in thumbnails:
            cid = f"img_{codigo.replace(' ', '_').replace('/', '_')}"
            imagen_html = f'<img src="cid:{cid}" style="width: 50px; height: 50px; object-fit: cover; border-radius: 4px;" alt="{codigo}">'
            imagenes_adjuntar[cid] = thumbnails[codigo]
        else:
            imagen_html = '<div style="width: 50px; height: 50px; background: #f0f0f0; border-radius: 4px; display: flex; align-items: center; justify-content: center; color: #999; font-size: 10px;">-</div>'

        filas_productos += f"""
        <tr>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{imagen_html}</td>
            <td style="padding: 12px; border: 1px solid #ddd; font-weight: bold; color: #667eea;">{item['codigo']}</td>
            <td style="padding: 12px; border: 1px solid #ddd;">{item['descripcion']}</td>
            <td style="padding: 12px; border: 1px solid #ddd; text-align: center;">{item.get('formato', '-')}</td>
            <td style="padding: 12px; border: 1px solid #ddd; text-align: center;">{item.get('calidad', '-')}</td>
            <td style="padding: 12px; border: 1px solid #ddd; text-align: center;">{item.get('tono', '-')}</td>
            <td style="padding: 12px; border: 1px solid #ddd; text-align: center;">{item.get('calibre', '-')}</td>
            <td style="padding: 12px; border: 1px solid #ddd; text-align: right;">{item['existencias']:,.2f} {item.get('unidad', '')}</td>
            <td style="padding: 12px; border: 1px solid #ddd; text-align: right; font-weight: bold; color: #667eea;">{item['cantidad_solicitada']:,.2f} {item.get('unidad', '')}</td>
        </tr>
        """

    # Sección de comentarios (solo si hay)
    comentarios_html = ""
    if comentarios:
        comentarios_html = f"""
        <div style="margin-top: 30px; padding: 15px; background-color: #e3f2fd; border-left: 4px solid #2196F3; border-radius: 4px;">
            <h3 style="margin-top: 0; color: #1976D2;">💬 Comentarios del Usuario:</h3>
            <p style="white-space: pre-wrap; margin: 0;">{comentarios}</p>
        </div>
        """

    # Sección de firma (solo si hay)
    firma_html = ""
    firma_cid = None
    if firma_base64:
        try:
            # Extraer datos base64 de la firma
            if firma_base64.startswith('data:'):
                firma_data = firma_base64.split(',')[1]
            else:
                firma_data = firma_base64

            firma_cid = "firma_digital"
            imagenes_adjuntar[firma_cid] = firma_data
            firma_html = f"""
            <div style="margin-top: 30px; padding: 15px; background-color: #f5f5f5; border: 1px solid #ddd; border-radius: 4px;">
                <h3 style="margin-top: 0; color: #333;">✍️ Firma Digital:</h3>
                <div style="background: white; padding: 10px; border: 1px dashed #ccc; border-radius: 4px; display: inline-block;">
                    <img src="cid:{firma_cid}" style="max-width: 250px; height: auto;" alt="Firma digital">
                </div>
            </div>
            """
            print(f"   ✅ Firma digital incluida en el email")
        except Exception as e:
            print(f"   ⚠️ Error procesando firma para email: {e}")

    # HTML completo del email
    body = f"""
    <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                          padding: 30px; text-align: center; color: white; border-radius: 10px 10px 0 0; }}
                .header h1 {{ margin: 0; font-size: 28px; }}
                .info-box {{ background-color: #f0f4ff; padding: 20px; margin: 20px 0; border-radius: 8px; }}
                .info-box p {{ margin: 8px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                     color: white; padding: 12px; text-align: left; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .footer {{ background-color: #fff9e6; padding: 15px; margin-top: 20px;
                          border-left: 4px solid #ffc107; border-radius: 4px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>📦 Solicitud de Stock</h1>
            </div>

            <div class="info-box">
                <p><strong>👤 Usuario:</strong> {usuario}</p>
                <p><strong>📅 Fecha:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
                {'<p><strong>🏢 Cliente:</strong> ' + cliente_info.get('codigo', '') + ' - ' + cliente_info.get('razon', '') + '</p>' if cliente_info else ''}
                {'<p><strong>🏷️ Referencia:</strong> ' + referencia + '</p>' if referencia else ''}
                <p><strong>📊 Total de productos:</strong> {len(carrito)}</p>
            </div>

            <h2 style="color: #667eea; border-bottom: 2px solid #667eea; padding-bottom: 10px;">
                📋 Detalle de Productos Solicitados
            </h2>

            <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                <thead>
                    <tr>
                        <th style="background-color: #667eea; color: white; padding: 12px; text-align: center; border: 1px solid #ddd; width: 60px;">Img</th>
                        <th style="background-color: #667eea; color: white; padding: 12px; text-align: left; border: 1px solid #ddd;">Código</th>
                        <th style="background-color: #667eea; color: white; padding: 12px; text-align: left; border: 1px solid #ddd;">Descripción</th>
                        <th style="background-color: #667eea; color: white; padding: 12px; text-align: center; border: 1px solid #ddd;">Formato</th>
                        <th style="background-color: #667eea; color: white; padding: 12px; text-align: center; border: 1px solid #ddd;">Calidad</th>
                        <th style="background-color: #667eea; color: white; padding: 12px; text-align: center; border: 1px solid #ddd;">Tono</th>
                        <th style="background-color: #667eea; color: white; padding: 12px; text-align: center; border: 1px solid #ddd;">Calibre</th>
                        <th style="background-color: #667eea; color: white; padding: 12px; text-align: right; border: 1px solid #ddd;">Stock Disponible</th>
                        <th style="background-color: #667eea; color: white; padding: 12px; text-align: right; border: 1px solid #ddd;">Cantidad Solicitada</th>
                    </tr>
                </thead>
                <tbody>
                    {filas_productos}
                </tbody>
            </table>

            {comentarios_html}

            {firma_html}

            <div class="footer">
                <p style="margin: 0;"><strong>📎 Nota:</strong> Se adjunta PDF y Excel con el detalle completo de esta solicitud.</p>
            </div>

            <p style="margin-top: 30px; color: #666;">
                <em>Este correo fue generado automáticamente por el Sistema de Gestión de Stocks.</em><br>
                <small>© {datetime.now().year} - Gestión de Inventarios</small>
            </p>
        </body>
    </html>
    """

    # Crear parte related para HTML + imágenes embebidas
    msg_related = MIMEMultipart('related')
    msg_related.attach(MIMEText(body, 'html'))

    # Adjuntar imágenes con CID
    if imagenes_adjuntar:
        print(f"   Adjuntando {len(imagenes_adjuntar)} imágenes con CID...")
        for cid, img_base64 in imagenes_adjuntar.items():
            try:
                img_data = base64.b64decode(img_base64)
                img_part = MIMEImage(img_data)
                img_part.add_header('Content-ID', f'<{cid}>')
                # La firma es PNG, el resto son JPG
                extension = '.png' if cid == 'firma_digital' else '.jpg'
                img_part.add_header('Content-Disposition', 'inline', filename=f'{cid}{extension}')
                msg_related.attach(img_part)
            except Exception as e:
                print(f"   ⚠️ Error adjuntando imagen {cid}: {e}")

    msg.attach(msg_related)
    print("✅ Cuerpo del email con tabla HTML e imágenes adjuntado")
    
    # Adjuntar PDF
    print("\n4️⃣ Adjuntando PDF...")
    part_pdf = MIMEBase('application', 'octet-stream')
    part_pdf.set_payload(pdf_buffer.read())
    encoders.encode_base64(part_pdf)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename_pdf = f'solicitud_stock_{timestamp}.pdf'
    part_pdf.add_header(
        'Content-Disposition',
        f'attachment; filename={filename_pdf}'
    )
    msg.attach(part_pdf)
    print(f"✅ PDF adjuntado: {filename_pdf}")

    # Adjuntar Excel
    if excel_buffer:
        print("\n4️⃣b Adjuntando Excel...")
        part_excel = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part_excel.set_payload(excel_buffer.read())
        encoders.encode_base64(part_excel)
        filename_excel = f'solicitud_stock_{timestamp}.xlsx'
        part_excel.add_header(
            'Content-Disposition',
            f'attachment; filename={filename_excel}'
        )
        msg.attach(part_excel)
        print(f"✅ Excel adjuntado: {filename_excel}")
    
    # Enviar email
    print("\n5️⃣ Conectando al servidor SMTP...")
    try:
        # Detectar si usar SSL o TLS según el puerto
        if email_config['smtp_port'] == 465:
            print(f"   Usando SSL en puerto {email_config['smtp_port']}")
            server = smtplib.SMTP_SSL(email_config['smtp_server'], email_config['smtp_port'])
        else:
            print(f"   Usando TLS en puerto {email_config['smtp_port']}")
            server = smtplib.SMTP(email_config['smtp_server'], email_config['smtp_port'])
            print("   Iniciando STARTTLS...")
            server.starttls()
        
        print("✅ Conexión SMTP establecida")
        
        print("\n6️⃣ Autenticando...")
        server.login(email_config['email_from'], email_config['email_password'])
        print("✅ Autenticación exitosa")
        
        print("\n7️⃣ Enviando mensaje...")
        server.send_message(msg)
        print("✅ Mensaje enviado correctamente")
        
        server.quit()
        print("\n✅ EMAIL ENVIADO EXITOSAMENTE")
        print("=" * 60 + "\n")
        
    except smtplib.SMTPAuthenticationError as e:
        print(f"\n❌ ERROR DE AUTENTICACIÓN SMTP")
        print(f"   Código: {e.smtp_code}")
        print(f"   Mensaje: {e.smtp_error}")
        print(f"   Verifica:")
        print(f"   - Usuario: {email_config['email_from']}")
        print(f"   - Contraseña correcta en la BD")
        print("=" * 60 + "\n")
        raise
    except smtplib.SMTPException as e:
        print(f"\n❌ ERROR SMTP")
        print(f"   {str(e)}")
        print("=" * 60 + "\n")
        raise
    except Exception as e:
        print(f"\n❌ ERROR INESPERADO AL ENVIAR EMAIL")
        print(f"   {str(e)}")
        traceback.print_exc()
        print("=" * 60 + "\n")
        raise