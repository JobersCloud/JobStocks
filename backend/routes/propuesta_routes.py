# ============================================================
#      ██╗ ██████╗ ██████╗ ███████╗██████╗ ███████╗
#      ██║██╔═══██╗██╔══██╗██╔════╝██╔══██╗██╔════╝
#      ██║██║   ██║██████╔╝█████╗  ██████╔╝███████╗
# ██   ██║██║   ██║██╔══██╗██╔══╝  ██╔══██╗╚════██║
# ╚█████╔╝╚██████╔╝██████╔╝███████╗██║  ██║███████║
#  ╚════╝  ╚═════╝ ╚═════╝ ╚══════╝╚═╝  ╚═╝╚══════╝
#
#                ──  Jobers - Iaucejo  ──
#
# Autor : iaucejo
# Fecha : 2026-01-08
# ============================================================

# ============================================
# ARCHIVO: routes/propuesta_routes.py
# Endpoints para consulta y gestion de propuestas (ERP)
# ============================================
from flask import Blueprint, jsonify, request, Response, session
from flask_login import login_required, current_user
from utils.auth import csrf_required
from utils.auth import api_key_or_login_required, administrador_required
from models.propuesta_model import PropuestaModel
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

propuesta_bp = Blueprint('propuesta', __name__, url_prefix='/api/propuestas')


@propuesta_bp.route('/mis-propuestas', methods=['GET'])
@login_required
def get_mis_propuestas():
    """
    Obtener las propuestas del usuario actual.
    Usa empresa de sesión.
    ---
    tags:
      - Propuestas
    security:
      - cookieAuth: []
    responses:
      200:
        description: Lista de propuestas del usuario
      401:
        description: No autenticado
    """
    # Usar empresa_erp de sesión para filtrar
    empresa_id = session.get('empresa_id', '1')

    try:
        propuestas = PropuestaModel.get_by_user(current_user.id, empresa_id=empresa_id)
        return jsonify({
            'success': True,
            'total': len(propuestas),
            'propuestas': propuestas
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@propuesta_bp.route('', methods=['GET'])
@login_required
@administrador_required
def get_todas_propuestas():
    """
    Obtener todas las propuestas (solo administradores).
    Usa empresa de sesión.
    ---
    tags:
      - Propuestas
    security:
      - cookieAuth: []
    parameters:
      - name: estado
        in: query
        type: string
        required: false
        enum: [Enviada, Procesada, Cancelada]
        description: Filtrar por estado
    responses:
      200:
        description: Lista de todas las propuestas
      401:
        description: No autenticado
      403:
        description: No autorizado (requiere rol administrador)
    """
    empresa_id = session.get('empresa_id', '1')
    estado = request.args.get('estado')

    try:
        propuestas = PropuestaModel.get_all(empresa_id=empresa_id, estado=estado)
        return jsonify({
            'success': True,
            'total': len(propuestas),
            'propuestas': propuestas
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@propuesta_bp.route('/pendientes', methods=['GET'])
@api_key_or_login_required
def get_pendientes():
    """
    Obtener propuestas pendientes de procesar.
    Usa empresa de sesión.
    ---
    tags:
      - Propuestas
    security:
      - cookieAuth: []
      - apiKeyHeader: []
      - apiKeyQuery: []
    parameters:
      - name: incluir_lineas
        in: query
        type: boolean
        required: false
        default: false
        description: Si es true, incluye las lineas de detalle
    responses:
      200:
        description: Lista de propuestas pendientes
      401:
        description: No autenticado
    """
    incluir_lineas = request.args.get('incluir_lineas', 'false').lower() == 'true'
    empresa_id = session.get('empresa_id', '1')

    try:
        propuestas = PropuestaModel.get_pendientes(incluir_lineas=incluir_lineas, empresa_id=empresa_id)
        return jsonify({
            'success': True,
            'total': len(propuestas),
            'propuestas': propuestas
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@propuesta_bp.route('/<int:propuesta_id>', methods=['GET'])
@api_key_or_login_required
def get_propuesta(propuesta_id):
    """
    Obtener una propuesta por ID con sus lineas
    ---
    tags:
      - Propuestas
    security:
      - cookieAuth: []
      - apiKeyHeader: []
      - apiKeyQuery: []
    parameters:
      - name: propuesta_id
        in: path
        type: integer
        required: true
        description: ID de la propuesta
    responses:
      200:
        description: Propuesta con sus lineas
      404:
        description: Propuesta no encontrada
      401:
        description: No autenticado
    """
    try:
        propuesta = PropuestaModel.get_by_id(propuesta_id)
        if not propuesta:
            return jsonify({
                'success': False,
                'error': 'Propuesta no encontrada'
            }), 404

        return jsonify({
            'success': True,
            'propuesta': propuesta
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@propuesta_bp.route('/lineas', methods=['GET'])
@api_key_or_login_required
def get_lineas_filtradas():
    """
    Obtener lineas de propuestas con filtro opcional
    ---
    tags:
      - Propuestas
    security:
      - cookieAuth: []
      - apiKeyHeader: []
      - apiKeyQuery: []
    parameters:
      - name: propuesta_id
        in: query
        type: integer
        required: false
        description: Filtrar por ID de propuesta
    responses:
      200:
        description: Lineas de propuestas
      401:
        description: No autenticado
    """
    propuesta_id = request.args.get('propuesta_id', type=int)

    try:
        lineas = PropuestaModel.get_lineas(propuesta_id=propuesta_id)
        return jsonify({
            'success': True,
            'total': len(lineas),
            'lineas': lineas
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@propuesta_bp.route('/<int:propuesta_id>/lineas', methods=['GET'])
@api_key_or_login_required
def get_lineas(propuesta_id):
    """
    Obtener las lineas de detalle de una propuesta
    ---
    tags:
      - Propuestas
    security:
      - cookieAuth: []
      - apiKeyHeader: []
      - apiKeyQuery: []
    parameters:
      - name: propuesta_id
        in: path
        type: integer
        required: true
        description: ID de la propuesta
    responses:
      200:
        description: Lineas de la propuesta
        schema:
          type: object
          properties:
            success:
              type: boolean
            propuesta_id:
              type: integer
            total:
              type: integer
            lineas:
              type: array
      404:
        description: Propuesta no encontrada
      401:
        description: No autenticado
    """
    try:
        propuesta = PropuestaModel.get_by_id(propuesta_id)
        if not propuesta:
            return jsonify({
                'success': False,
                'error': 'Propuesta no encontrada'
            }), 404

        return jsonify({
            'success': True,
            'propuesta_id': propuesta_id,
            'total': len(propuesta['lineas']),
            'lineas': propuesta['lineas']
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@propuesta_bp.route('/<int:propuesta_id>/estado', methods=['PUT'])
@api_key_or_login_required
@csrf_required
def actualizar_estado(propuesta_id):
    """
    Actualizar el estado de una propuesta
    ---
    tags:
      - Propuestas
    security:
      - cookieAuth: []
      - apiKeyHeader: []
      - apiKeyQuery: []
    parameters:
      - name: propuesta_id
        in: path
        type: integer
        required: true
        description: ID de la propuesta
      - name: body
        in: body
        required: true
        schema:
          type: object
          required:
            - estado
          properties:
            estado:
              type: string
              enum: [Enviada, Procesada, Cancelada]
              description: Nuevo estado de la propuesta
    responses:
      200:
        description: Estado actualizado correctamente
      400:
        description: Estado invalido
      404:
        description: Propuesta no encontrada
      401:
        description: No autenticado
    """
    data = request.json

    if not data or 'estado' not in data:
        return jsonify({
            'success': False,
            'error': 'Campo "estado" es requerido'
        }), 400

    estado = data['estado']

    try:
        actualizado = PropuestaModel.actualizar_estado(propuesta_id, estado)
        if not actualizado:
            return jsonify({
                'success': False,
                'error': 'Propuesta no encontrada'
            }), 404

        # Notificar al usuario propietario del cambio de estado
        try:
            from models.notification_model import NotificationModel
            propuesta = PropuestaModel.get_by_id(propuesta_id)
            if propuesta and propuesta.get('user_id'):
                empresa_id = session.get('empresa_id')
                connection_id = session.get('connection')
                NotificationModel.create(
                    user_id=propuesta['user_id'],
                    empresa_id=empresa_id,
                    tipo='propuesta_cambio_estado',
                    titulo=f'Propuesta #{propuesta_id} cambió a: {estado}',
                    mensaje=f'Tu propuesta #{propuesta_id} ha cambiado al estado "{estado}".',
                    data={'propuesta_id': propuesta_id, 'estado': estado},
                    connection_id=connection_id
                )
        except Exception as e:
            print(f"Warning: No se pudo crear notificación: {e}")

        return jsonify({
            'success': True,
            'message': f'Estado actualizado a "{estado}"',
            'propuesta_id': propuesta_id,
            'nuevo_estado': estado
        }), 200
    except ValueError as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@propuesta_bp.route('/<int:propuesta_id>/excel', methods=['GET'])
@login_required
def descargar_excel(propuesta_id):
    """
    Descargar Excel de una propuesta
    ---
    tags:
      - Propuestas
    security:
      - cookieAuth: []
    parameters:
      - name: propuesta_id
        in: path
        type: integer
        required: true
        description: ID de la propuesta
    responses:
      200:
        description: Archivo Excel de la propuesta
      404:
        description: Propuesta no encontrada
      403:
        description: No autorizado para ver esta propuesta
      401:
        description: No autenticado
    """
    try:
        propuesta = PropuestaModel.get_by_id(propuesta_id)
        if not propuesta:
            return jsonify({
                'success': False,
                'error': 'Propuesta no encontrada'
            }), 404

        # Verificar que el usuario puede ver esta propuesta
        # (es suya o es administrador)
        if propuesta['user_id'] != current_user.id and not current_user.is_administrador():
            return jsonify({
                'success': False,
                'error': 'No autorizado para ver esta propuesta'
            }), 403

        # Generar Excel
        excel_buffer = generar_excel_propuesta(propuesta)

        # Preparar respuesta
        filename = f"propuesta_{propuesta_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return Response(
            excel_buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}'
            }
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def generar_excel_propuesta(propuesta):
    """Genera un archivo Excel con los datos de una propuesta"""
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Propuesta"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    number_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Información de cabecera
    ws['A1'] = "PROPUESTA DE STOCK"
    ws['A1'].font = Font(bold=True, size=14, color="667eea")
    ws.merge_cells('A1:J1')

    ws['A2'] = f"Propuesta ID: {propuesta['id']}"
    ws['A3'] = f"Usuario: {propuesta.get('full_name', propuesta.get('username', 'N/A'))}"
    ws['A4'] = f"Fecha: {propuesta.get('fecha', '')}"
    ws['A5'] = f"Estado: {propuesta.get('estado', '')}"

    start_row = 7

    # Encabezados de la tabla
    headers = ['Código', 'Descripción', 'Formato', 'Calidad', 'Tono', 'Calibre', 'Unidad', 'Pallet', 'Caja', 'Cantidad']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos de las líneas
    lineas = propuesta.get('lineas', [])
    for row_num, item in enumerate(lineas, start_row + 1):
        ws.cell(row=row_num, column=1, value=item.get('codigo', '')).border = thin_border
        ws.cell(row=row_num, column=2, value=item.get('descripcion', '')).border = thin_border
        ws.cell(row=row_num, column=3, value=item.get('formato', '')).border = thin_border
        ws.cell(row=row_num, column=4, value=item.get('calidad', '')).border = thin_border
        ws.cell(row=row_num, column=5, value=item.get('tono', '')).border = thin_border
        ws.cell(row=row_num, column=6, value=item.get('calibre', '')).border = thin_border
        ws.cell(row=row_num, column=7, value=item.get('unidad', '')).border = thin_border
        ws.cell(row=row_num, column=8, value=item.get('pallet', '')).border = thin_border
        ws.cell(row=row_num, column=9, value=item.get('caja', '')).border = thin_border

        cantidad_cell = ws.cell(row=row_num, column=10, value=item.get('cantidad_solicitada', 0))
        cantidad_cell.border = thin_border
        cantidad_cell.alignment = number_alignment

    # Comentarios al final
    comentarios = propuesta.get('comentarios', '')
    if comentarios:
        last_row = start_row + len(lineas) + 2
        ws.cell(row=last_row, column=1, value="Comentarios:").font = Font(bold=True)
        ws.cell(row=last_row + 1, column=1, value=comentarios)
        ws.merge_cells(f'A{last_row + 1}:J{last_row + 1}')

    # Ajustar anchos de columna
    column_widths = [15, 40, 12, 10, 8, 10, 8, 10, 10, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    wb.save(buffer)
    buffer.seek(0)
    return buffer
